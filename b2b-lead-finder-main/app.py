"""
B2B Lead Finder – Streamlit App
================================
Run:  streamlit run app.py
"""

import sys
import uuid
import json
import subprocess
import datetime
import time
import pandas as pd
from pathlib import Path

import streamlit as st

# ─── Install Playwright browser on first boot (needed on Streamlit Cloud) ────
@st.cache_resource
def _install_playwright_browser():
    result = subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        capture_output=True, text=True
    )
    return result.returncode

_install_playwright_browser()

# ─── Config ───────────────────────────────────────────────────────────────────
HISTORY_FILE = Path(__file__).parent / "history.json"
RESULTS_DIR  = Path(__file__).parent / "results"

LOCATIONS = {
    "Singapore": [
        "Singapore", "Central Region", "East Region", "North Region",
        "North-East Region", "West Region", "Orchard", "Marina Bay",
        "Tanjong Pagar", "Bugis", "Jurong", "Tampines", "Woodlands",
        "Ang Mo Kio", "Bishan", "Clementi", "Bedok", "Pasir Ris",
        "Punggol", "Sengkang", "Hougang", "Serangoon", "Buona Vista",
        "Novena", "Toa Payoh",
    ],
    "Malaysia": [
        "Malaysia", "Kuala Lumpur", "Selangor", "Johor", "Penang",
        "Perak", "Sabah", "Sarawak", "Pahang", "Negeri Sembilan",
        "Melaka", "Kedah", "Kelantan", "Terengganu", "Perlis",
        "Putrajaya", "Labuan",
    ],
}

# ─── History ──────────────────────────────────────────────────────────────────
def load_history():
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []

def save_history(query, country, leads):
    history = load_history()
    entry = {
        "query"    : query,
        "country"  : country,
        "count"    : len(leads),
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "leads"    : leads,
    }
    for i, h in enumerate(history):
        if h["query"].lower() == query.lower() and h["country"] == country:
            history[i] = entry
            HISTORY_FILE.write_text(json.dumps(history, indent=2, ensure_ascii=False), encoding="utf-8")
            return
    history.insert(0, entry)
    HISTORY_FILE.write_text(json.dumps(history[:30], indent=2, ensure_ascii=False), encoding="utf-8")

# ─── AI Commentary ────────────────────────────────────────────────────────────
def ai_msg(text, ts=None):
    """Append a message to the AI feed in session state."""
    if "ai_feed" not in st.session_state:
        st.session_state.ai_feed = []
    st.session_state.ai_feed.append({
        "text": text,
        "time": ts or datetime.datetime.now().strftime("%H:%M:%S"),
    })

def ai_insights(leads, query, location):
    """Generate smart commentary from the scraped leads data."""
    n = len(leads)
    if n == 0:
        return f"Searched for **{query}** in **{location}** but came up empty. Google Maps might not have many listings for that category there — try a broader search term."

    df = pd.DataFrame(leads)

    has_phone   = int(df["Phone"].astype(str).str.len().gt(2).sum())
    has_website = int(df["Website"].astype(str).str.len().gt(4).sum())
    has_both    = int((df["Phone"].astype(str).str.len().gt(2) & df["Website"].astype(str).str.len().gt(4)).sum())
    no_contact  = n - int((df["Phone"].astype(str).str.len().gt(2) | df["Website"].astype(str).str.len().gt(4)).sum())

    # Top rated
    rated = df[df["Rating"].astype(str).str.match(r"^\d")]
    top_name, top_rating, top_reviews = "", "", 0
    if not rated.empty:
        rated = rated.copy()
        rated["_r"] = pd.to_numeric(rated["Rating"], errors="coerce")
        best = rated.sort_values("_r", ascending=False).iloc[0]
        top_name    = best["Name"]
        top_rating  = best["Rating"]
        top_reviews = best.get("Reviews", 0)

    # Most reviewed
    most_reviewed_name, most_reviews = "", 0
    if "Reviews" in df.columns:
        mx = df.loc[pd.to_numeric(df["Reviews"], errors="coerce").fillna(0).idxmax()]
        most_reviewed_name = mx["Name"]
        most_reviews       = int(mx["Reviews"]) if str(mx["Reviews"]).isdigit() else 0

    lines = []
    lines.append(f"✅ Scraped **{n} {query}** in **{location}**.")
    lines.append(f"📞 **{has_phone}** have a phone number · 🌐 **{has_website}** have a website · 📞🌐 **{has_both}** have both.")
    if no_contact > 0:
        lines.append(f"🎯 **{no_contact}** businesses have no online presence — prime cold outreach targets.")
    if top_name:
        lines.append(f"⭐ Top rated: **{top_name}** ({top_rating}★{f', {top_reviews} reviews' if top_reviews else ''})")
    if most_reviewed_name and most_reviews > 0:
        lines.append(f"💬 Most reviewed: **{most_reviewed_name}** with **{most_reviews}** reviews.")
    phone_pct = round(has_phone / n * 100)
    web_pct   = round(has_website / n * 100)
    lines.append(f"📊 Contact coverage: {phone_pct}% have phones, {web_pct}% have websites.")

    return "\n\n".join(lines)

# ─── Excel helper ─────────────────────────────────────────────────────────────
@st.cache_data
def to_excel(records):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb  = openpyxl.Workbook(); ws = wb.active; ws.title = "Leads"
    hdrs  = list(records[0].keys())
    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1A1A2E")
    thin  = Side(style="thin", color="DDDDDD")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font=hf; c.fill=hfill; c.border=bdr
        c.alignment=Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(records, 2):
        for ci, key in enumerate(hdrs, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(key,""))
            c.border=bdr; c.alignment=Alignment(vertical="center")
            if key=="Google Maps" and row.get(key):
                c.hyperlink=row[key]; c.font=Font(color="4361EE", underline="single")
            elif key=="Name":
                c.font=Font(bold=True)
    for i in range(1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(i)].width = 28
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def read_job(job_id):
    job_file = RESULTS_DIR / f"{job_id}.json"
    if not job_file.exists():
        return None
    try:
        return json.loads(job_file.read_text(encoding="utf-8"))
    except Exception:
        return None

# ─── Streamlit UI ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="B2B Lead Finder", page_icon="🔍", layout="wide")
st.markdown("""
<style>
  [data-testid="stSidebar"] { background: #161628; }
  .block-container { padding-top: 1.5rem; }
  .ai-bubble {
      background: #1e1e38;
      border-left: 3px solid #4361EE;
      border-radius: 6px;
      padding: 10px 12px;
      margin-bottom: 10px;
      font-size: 0.82rem;
      line-height: 1.5;
      color: #c9cfe8;
  }
  .ai-time {
      font-size: 0.7rem;
      color: #555a7a;
      margin-bottom: 4px;
  }
  .ai-header {
      color: #4361EE;
      font-weight: 700;
      font-size: 0.78rem;
      letter-spacing: 0.05em;
      text-transform: uppercase;
      margin-bottom: 8px;
  }
</style>""", unsafe_allow_html=True)

# Session state defaults
for k, v in {
    "leads"     : [],
    "status"    : "",
    "job_id"    : None,
    "running"   : False,
    "proc_pid"  : None,
    "total"     : 0,
    "ai_feed"   : [],
    "last_n"    : 0,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔍 B2B Lead Finder")
    st.caption("Google Maps · Live Scraper")
    st.divider()

    st.markdown("**Previous Searches**")
    history = load_history()
    if not history:
        st.caption("No searches yet")
    else:
        for idx, h in enumerate(history):
            if st.button(f"{h['query']} · {h['country']} · {h['count']} leads",
                         key=f"h_{idx}", use_container_width=True):
                st.session_state.leads   = h["leads"]
                st.session_state.running = False
                st.session_state.job_id  = None
                st.session_state.status  = f"✅ Loaded: **{h['query']}** in {h['country']} — {h['count']} leads"
                st.rerun()

    # ── AI Feed ───────────────────────────────────────────────────────────────
    if st.session_state.ai_feed:
        st.divider()
        st.markdown('<div class="ai-header">🤖 AI Assistant</div>', unsafe_allow_html=True)
        for msg in reversed(st.session_state.ai_feed[-6:]):  # show latest 6
            st.markdown(
                f'<div class="ai-bubble">'
                f'<div class="ai-time">{msg["time"]}</div>'
                f'{msg["text"]}'
                f'</div>',
                unsafe_allow_html=True,
            )

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("## B2B Lead Finder")
st.caption("Find business leads from Google Maps — leads appear live as they're found")

c1, c2, c3, c4, c5 = st.columns([3, 1.5, 2, 1, 1])
with c1:
    query = st.text_input("Search", placeholder="e.g. interior designers, dentists, lawyers …", label_visibility="collapsed")
with c2:
    country = st.selectbox("Country", list(LOCATIONS.keys()), label_visibility="collapsed")
with c3:
    location = st.selectbox("State / Region", LOCATIONS[country], label_visibility="collapsed")
with c4:
    start_btn = st.button("🚀 Start", use_container_width=True, type="primary",
                          disabled=st.session_state.running)
with c5:
    stop_btn = st.button("⏹ Stop", use_container_width=True,
                         disabled=not st.session_state.running)

status_ph   = st.empty()
progress_ph = st.empty()
table_ph    = st.empty()

# ── Stop ──────────────────────────────────────────────────────────────────────
if stop_btn and st.session_state.running:
    if st.session_state.proc_pid:
        try:
            import os, signal
            os.kill(st.session_state.proc_pid, signal.SIGTERM)
        except Exception:
            pass
    st.session_state.running = False
    st.session_state.status  = "⏹ Stopped."
    n = len(st.session_state.leads)
    if n:
        ai_msg(f"⏹ Search stopped early. Collected **{n}** leads so far.")
    st.rerun()

# ── Start ─────────────────────────────────────────────────────────────────────
if start_btn and query and not st.session_state.running:
    job_id = str(uuid.uuid4())
    RESULTS_DIR.mkdir(exist_ok=True)

    worker = Path(__file__).parent / "scraper_worker.py"
    proc = subprocess.Popen(
        [sys.executable, str(worker), query, location, job_id],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    st.session_state.leads    = []
    st.session_state.job_id   = job_id
    st.session_state.running  = True
    st.session_state.proc_pid = proc.pid
    st.session_state.status   = "🌐 Starting scraper …"
    st.session_state.total    = 80
    st.session_state.last_n   = 0

    ai_msg(f"🔍 Starting search for **{query}** in **{location}**. Opening Google Maps and waiting for results…")
    st.rerun()

# ── Poll running job ──────────────────────────────────────────────────────────
if st.session_state.running and st.session_state.job_id:
    job_id = st.session_state.job_id
    data   = read_job(job_id)

    if data is None:
        status_ph.info("🌐 Starting scraper …")
        time.sleep(1)
        st.rerun()
    else:
        leads   = data.get("leads", [])
        message = data.get("message", "")
        total   = data.get("total", st.session_state.total) or 80
        done    = data.get("status") == "done"

        st.session_state.leads = leads
        st.session_state.total = total

        n   = len(leads)
        pct = min(n / total, 1.0) if total else 0

        # Milestone AI messages during scraping
        prev_n = st.session_state.last_n
        if total > 0 and prev_n == 0 and n == 0 and "listings" in message.lower():
            ai_msg(f"📌 Found **{total} listings** on Google Maps. Starting to extract details for each one…")
        if prev_n < 10 <= n:
            ai_msg(f"⚡ First 10 leads extracted! Looking good so far — **{total - n}** more to go.")
        if prev_n < 25 <= n:
            ai_msg(f"📈 25 leads in. Scraper is running smooth — keeping at it.")
        if prev_n < 50 <= n:
            ai_msg(f"🔥 50 leads and counting. This is a busy area for **{query}**!")
        if prev_n < 75 <= n:
            ai_msg(f"💪 75 leads extracted. Almost there!")
        st.session_state.last_n = n

        if done:
            st.session_state.running = False
            st.session_state.job_id  = None
            progress_ph.empty()
            is_error = message.startswith("❌") or message.startswith("⚠️")
            if n > 0:
                status_ph.success(f"✅ Done! Found **{n}** leads for **{query}** in **{location}**")
                save_history(query, location, leads)
                ai_msg(ai_insights(leads, query, location))
            elif is_error:
                status_ph.error(f"{message}")
                ai_msg(f"⚠️ Search ended with an issue: {message}")
            else:
                status_ph.warning(f"⚠️ Done but found **0 leads** for **{query}** in **{location}**. Last status: _{message}_")
                ai_msg(f"🤔 Came back empty for **{query}** in **{location}**. Last status from scraper: _{message}_")
        else:
            status_ph.info(message or f"⚡ Scraping … **{n}** leads found so far")
            if n > 0:
                progress_ph.progress(pct, text=f"{n} / {total} processed")

        if leads:
            table_ph.dataframe(
                pd.DataFrame(leads),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Google Maps": st.column_config.LinkColumn("Maps 📍"),
                    "Website"    : st.column_config.LinkColumn("Website 🌐"),
                    "Rating"     : st.column_config.NumberColumn("Rating ⭐", format="%.1f"),
                    "Reviews"    : st.column_config.NumberColumn("Reviews 💬"),
                },
            )

        if not done:
            time.sleep(2)
            st.rerun()

# ── Display existing leads (not running) ─────────────────────────────────────
elif st.session_state.leads and not st.session_state.running:
    if st.session_state.status:
        status_ph.info(st.session_state.status)

    df = pd.DataFrame(st.session_state.leads)

    f1, f2, f3 = st.columns([3, 2, 2])
    with f1:
        sf = st.text_input("Filter", placeholder="Search name, address, phone …", label_visibility="collapsed")
    with f2:
        cf = st.selectbox("Contact", ["All leads", "Has phone", "Has website", "Phone + website"], label_visibility="collapsed")
    with f3:
        so = st.selectbox("Sort", ["Reviews ↓", "Reviews ↑", "Rating ↓", "Name A–Z"], label_visibility="collapsed")

    if sf:
        mask = df.apply(lambda r: sf.lower() in r.astype(str).str.lower().str.cat(), axis=1)
        df   = df[mask]
    if cf == "Has phone":
        df = df[df["Phone"].astype(str).str.len() > 2]
    elif cf == "Has website":
        df = df[df["Website"].astype(str).str.len() > 4]
    elif cf == "Phone + website":
        df = df[(df["Phone"].astype(str).str.len() > 2) & (df["Website"].astype(str).str.len() > 4)]

    sc, asc = {"Reviews ↓":("Reviews",False),"Reviews ↑":("Reviews",True),"Rating ↓":("Rating",False),"Name A–Z":("Name",True)}[so]
    try:
        df = df.sort_values(sc, ascending=asc)
    except Exception:
        pass

    st.caption(f"Showing **{len(df)}** of **{len(st.session_state.leads)}** leads")
    st.dataframe(
        df, use_container_width=True, hide_index=True,
        column_config={
            "Google Maps": st.column_config.LinkColumn("Maps 📍"),
            "Website"    : st.column_config.LinkColumn("Website 🌐"),
            "Rating"     : st.column_config.NumberColumn("Rating ⭐", format="%.1f ⭐"),
            "Reviews"    : st.column_config.NumberColumn("Reviews 💬"),
        },
    )
    st.download_button(
        "⬇️ Download Excel",
        data=to_excel(tuple(st.session_state.leads)),
        file_name=f"leads_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.markdown("""
    <div style="text-align:center;padding:60px 0;color:#7c829e">
      <div style="font-size:3rem">🏢</div>
      <div style="margin-top:12px;font-size:1.1rem">Search for any type of business in any country</div>
      <div style="margin-top:8px;font-size:.85rem">Leads appear live as they're scraped from Google Maps</div>
    </div>""", unsafe_allow_html=True)
