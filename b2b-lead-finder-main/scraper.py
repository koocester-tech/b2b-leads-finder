"""
Singapore Interior Design – Google Maps Lead Scraper
=====================================================
Scrapes Google Maps for interior design businesses in Singapore.
No login required. Just run and get leads.

Collects: business name, rating, review count, address,
          phone number, website, Google Maps URL

Output:
  sg_leads_TIMESTAMP.xlsx
  sg_leads_dashboard_TIMESTAMP.html

Usage:
  python scraper.py
"""

import re
import json
import time
import datetime
from pathlib import Path

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("❌  playwright not found. Run:  python -m pip install playwright")
    print("    Then run:                   python -m playwright install chromium")
    raise SystemExit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌  openpyxl not found. Run:  python -m pip install openpyxl")
    raise SystemExit(1)

# ─── Config ───────────────────────────────────────────────────────────────────
SEARCH_QUERIES = [
    "interior design Singapore",
    "interior designer Singapore",
    "home renovation Singapore",
    "ID firm Singapore",
]

MAX_RESULTS_PER_QUERY = 60      # Google Maps shows ~60-80 per search
OUTPUT_DIR = Path(".")

# ─── Helpers ──────────────────────────────────────────────────────────────────

def clean(text):
    return (text or "").strip()

def scroll_panel(page, panel_selector, times=15):
    for _ in range(times):
        page.evaluate("""(sel) => {
            const el = document.querySelector(sel);
            if (el) el.scrollTop += 800;
        }""", panel_selector)
        time.sleep(0.8)

# ─── Scraper ──────────────────────────────────────────────────────────────────

def scrape_google_maps(query: str, max_results: int) -> list[dict]:
    leads = []
    seen_names = set()

    with sync_playwright() as p:
        print(f"\n🔍  Searching: \"{query}\"")

        browser = p.chromium.launch(headless=False)   # headless=False so you can watch it
        context = browser.new_context(
            viewport={"width": 1280, "height": 800},
            locale="en-US",
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}/"
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        time.sleep(4)

        # dismiss cookie banner if present
        try:
            page.click("button:has-text('Accept all')", timeout=3000)
        except Exception:
            pass
        try:
            page.click("button:has-text('Reject all')", timeout=3000)
        except Exception:
            pass

        # scroll the results panel to load more listings
        panel_sel = "div[role='feed']"
        print("   📜  Scrolling to load results …")
        scroll_panel(page, panel_sel, times=20)

        # collect all listing elements
        cards = page.query_selector_all("div[role='feed'] > div > div > a")
        print(f"   📌  Found {len(cards)} listings")

        for i, card in enumerate(cards[:max_results]):
            try:
                card.click()
                time.sleep(1.8)

                # ── name ──────────────────────────────────────────────────
                name_el = page.query_selector("h1.DUwDvf, h1[class*='fontHeadline']")
                name = clean(name_el.inner_text()) if name_el else ""
                if not name or name in seen_names:
                    continue
                seen_names.add(name)

                # ── rating & reviews ──────────────────────────────────────
                rating_el = page.query_selector("div.F7nice span[aria-hidden='true']")
                rating = clean(rating_el.inner_text()) if rating_el else ""

                reviews_el = page.query_selector("div.F7nice span[aria-label*='review']")
                reviews_text = clean(reviews_el.get_attribute("aria-label")) if reviews_el else ""
                reviews_match = re.search(r"([\d,]+)", reviews_text)
                reviews = reviews_match.group(1).replace(",", "") if reviews_match else "0"

                # ── category ──────────────────────────────────────────────
                cat_el = page.query_selector("button.DkEaL, span.mgr77e")
                category = clean(cat_el.inner_text()) if cat_el else ""

                # ── address ───────────────────────────────────────────────
                addr_el = page.query_selector("button[data-item-id='address'] div.rogA2c")
                address = clean(addr_el.inner_text()) if addr_el else ""

                # ── phone ─────────────────────────────────────────────────
                phone_el = page.query_selector("button[data-item-id*='phone'] div.rogA2c")
                phone = clean(phone_el.inner_text()) if phone_el else ""

                # ── website ───────────────────────────────────────────────
                web_el = page.query_selector("a[data-item-id='authority']")
                website = clean(web_el.get_attribute("href")) if web_el else ""
                website = re.sub(r"\?.*", "", website)   # strip query params

                # ── Maps URL ──────────────────────────────────────────────
                maps_url = page.url.split("?")[0]

                lead = {
                    "name"       : name,
                    "rating"     : rating,
                    "reviews"    : int(reviews) if reviews.isdigit() else 0,
                    "category"   : category,
                    "address"    : address,
                    "phone"      : phone,
                    "website"    : website,
                    "maps_url"   : maps_url,
                    "scraped_at" : datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                }
                leads.append(lead)
                status = "⭐" if float(rating) >= 4.5 else "✅" if rating else "✅"
                print(f"   {status}  {name[:45]:<45} | {rating or '—'} ⭐ | {reviews} reviews")

            except Exception as e:
                continue

        browser.close()

    return leads


def scrape_all() -> list[dict]:
    all_leads = []
    seen = set()

    for query in SEARCH_QUERIES:
        results = scrape_google_maps(query, MAX_RESULTS_PER_QUERY)
        for lead in results:
            if lead["name"] not in seen:
                seen.add(lead["name"])
                all_leads.append(lead)

    print(f"\n🎯  Total unique leads: {len(all_leads)}\n")
    return all_leads


# ─── Excel export ─────────────────────────────────────────────────────────────

HEADERS = ["Business Name", "Rating", "Reviews", "Category",
           "Phone", "Website", "Address", "Google Maps", "Scraped At"]
FIELD_MAP = ["name", "rating", "reviews", "category",
             "phone", "website", "address", "maps_url", "scraped_at"]

HEADER_COLOR = "1A1A2E"
ROW1 = "FFFFFF"
ROW2 = "F4F6FF"
ACCENT = "4361EE"


def export_excel(leads, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SG Interior Design Leads"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=HEADER_COLOR)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = bdr
    ws.row_dimensions[1].height = 28

    for ri, lead in enumerate(leads, 2):
        fc = ROW1 if ri % 2 == 0 else ROW2
        rf = PatternFill("solid", fgColor=fc)
        for ci, field in enumerate(FIELD_MAP, 1):
            val = lead.get(field, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = rf; c.border = bdr
            c.alignment = Alignment(vertical="center")
            if field == "maps_url" and val:
                c.hyperlink = val
                c.font = Font(color=ACCENT, underline="single")
            elif field == "name":
                c.font = Font(bold=True)

    for i, w in enumerate([32, 8, 10, 22, 18, 35, 40, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"📊  Excel saved → {path}")


# ─── HTML Dashboard ───────────────────────────────────────────────────────────

def export_dashboard(leads, path):
    leads_json = json.dumps(leads, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SG Interior Design Leads</title>
<style>
  :root {{
    --bg:#0f0f1a;--surf:#1a1a2e;--surf2:#16213e;
    --acc:#4361ee;--acc2:#7209b7;--grn:#06d6a0;--yel:#ffd166;
    --txt:#e2e8f0;--mut:#94a3b8;--bdr:#2d3748;--rad:10px;
  }}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:var(--bg);color:var(--txt);font-family:'Segoe UI',system-ui,sans-serif}}
  .topbar{{background:var(--surf);border-bottom:1px solid var(--bdr);padding:16px 28px;display:flex;align-items:center;gap:14px;flex-wrap:wrap}}
  .topbar h1{{font-size:1.2rem;font-weight:700;background:linear-gradient(135deg,var(--acc),var(--acc2));-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
  .badge{{background:var(--acc);color:#fff;border-radius:20px;padding:3px 12px;font-size:.78rem;font-weight:600}}
  .stats{{display:flex;gap:12px;padding:20px 28px 0;flex-wrap:wrap}}
  .sc{{background:var(--surf);border:1px solid var(--bdr);border-radius:var(--rad);padding:14px 20px;min-width:130px}}
  .sc .v{{font-size:1.8rem;font-weight:800;color:var(--acc)}}
  .sc .l{{font-size:.7rem;color:var(--mut);margin-top:3px;text-transform:uppercase;letter-spacing:.05em}}
  .ctrls{{display:flex;gap:10px;padding:18px 28px;flex-wrap:wrap;align-items:center}}
  input[type=text],select{{background:var(--surf);border:1px solid var(--bdr);color:var(--txt);border-radius:8px;padding:7px 12px;font-size:.85rem;outline:none}}
  input[type=text]{{width:220px}}
  .btn{{background:var(--surf);border:1px solid var(--bdr);color:var(--txt);border-radius:8px;padding:7px 16px;font-size:.85rem;cursor:pointer;font-weight:600}}
  .tw{{padding:0 28px 32px;overflow-x:auto}}
  table{{width:100%;border-collapse:collapse;font-size:.84rem}}
  thead th{{background:var(--surf2);color:var(--mut);font-weight:600;text-transform:uppercase;letter-spacing:.06em;font-size:.7rem;padding:11px 13px;border-bottom:2px solid var(--bdr);white-space:nowrap;cursor:pointer}}
  thead th:hover{{color:var(--txt)}}
  tbody tr{{border-bottom:1px solid var(--bdr);transition:background .1s}}
  tbody tr:hover{{background:var(--surf)}}
  td{{padding:10px 13px;vertical-align:middle}}
  .pill{{display:inline-block;border-radius:20px;padding:2px 10px;font-size:.72rem;font-weight:600;white-space:nowrap}}
  .g{{background:rgba(6,214,160,.15);color:var(--grn)}}
  .y{{background:rgba(255,209,102,.15);color:var(--yel)}}
  .b{{background:rgba(67,97,238,.15);color:#a5b4fc}}
  a{{color:var(--acc);text-decoration:none}}
  a:hover{{text-decoration:underline}}
  .stars{{color:var(--yel)}}
</style>
</head>
<body>
<div class="topbar">
  <h1>🏠 SG Interior Design Leads</h1>
  <span class="badge" id="badge">0 leads</span>
  <span style="margin-left:auto;color:var(--mut);font-size:.78rem">Source: Google Maps Singapore</span>
</div>
<div class="stats" id="stats"></div>
<div class="ctrls">
  <input type="text" id="search" placeholder="Search name, address, phone…" oninput="render()">
  <select id="sc" onchange="render()">
    <option value="reviews">Sort: Reviews</option>
    <option value="rating">Sort: Rating</option>
    <option value="name">Sort: Name</option>
  </select>
  <select id="sd" onchange="render()">
    <option value="desc">High → Low</option>
    <option value="asc">Low → High</option>
  </select>
  <select id="fi" onchange="render()">
    <option value="all">All leads</option>
    <option value="phone">Has phone</option>
    <option value="website">Has website</option>
    <option value="both">Has phone + website</option>
  </select>
  <button class="btn" onclick="exportCSV()">⬇ Export CSV</button>
</div>
<div class="tw">
  <table>
    <thead><tr>
      <th>#</th><th>Business Name</th><th>Rating</th><th>Reviews</th>
      <th>Category</th><th>Phone</th><th>Website</th><th>Address</th><th>Maps</th>
    </tr></thead>
    <tbody id="tb"></tbody>
  </table>
</div>
<script>
const DATA={leads_json};
let filtered=[...DATA];

function stars(r){{
  if(!r)return'<span style="color:var(--mut)">—</span>';
  const n=parseFloat(r);
  return`<span class="pill ${{n>=4.5?'g':n>=4?'y':'b'}}">★ ${{r}}</span>`;
}}
function fmt(n){{return n>=1000?(n/1000).toFixed(1)+'k':n.toString()}}

function renderStats(){{
  const t=DATA.length;
  const wp=DATA.filter(l=>l.phone).length;
  const ww=DATA.filter(l=>l.website).length;
  const avgR=DATA.filter(l=>l.rating).length;
  const topR=DATA.filter(l=>parseFloat(l.rating)>=4.5).length;
  document.getElementById('badge').textContent=t+' leads';
  document.getElementById('stats').innerHTML=`
    <div class="sc"><div class="v">${{t}}</div><div class="l">Total Leads</div></div>
    <div class="sc"><div class="v">${{wp}}</div><div class="l">Have Phone</div></div>
    <div class="sc"><div class="v">${{ww}}</div><div class="l">Have Website</div></div>
    <div class="sc"><div class="v">${{topR}}</div><div class="l">Rating ≥ 4.5★</div></div>
  `;
}}

function render(){{
  const q=document.getElementById('search').value.toLowerCase();
  const sc=document.getElementById('sc').value;
  const sd=document.getElementById('sd').value;
  const fi=document.getElementById('fi').value;
  filtered=DATA.filter(l=>{{
    const t=[l.name,l.address,l.phone,l.website,l.category].join(' ').toLowerCase();
    if(q&&!t.includes(q))return false;
    if(fi==='phone'&&!l.phone)return false;
    if(fi==='website'&&!l.website)return false;
    if(fi==='both'&&(!l.phone||!l.website))return false;
    return true;
  }});
  filtered.sort((a,b)=>{{
    let av=a[sc]??'',bv=b[sc]??'';
    if(typeof av==='string'){{av=av.toLowerCase();bv=bv.toLowerCase()}}
    return sd==='asc'?(av>bv?1:-1):(av<bv?1:-1);
  }});
  document.getElementById('tb').innerHTML=filtered.map((l,i)=>`
    <tr>
      <td style="color:var(--mut)">${{i+1}}</td>
      <td style="font-weight:700">${{l.name}}</td>
      <td>${{stars(l.rating)}}</td>
      <td><span class="pill b">${{fmt(l.reviews)}}</span></td>
      <td style="color:var(--mut);font-size:.78rem">${{l.category||'—'}}</td>
      <td>${{l.phone?`<a href="tel:${{l.phone}}">${{l.phone}}</a>`:'<span style="color:var(--mut)">—</span>'}}</td>
      <td>${{l.website?`<a href="${{l.website}}" target="_blank">🌐 visit</a>`:'<span style="color:var(--mut)">—</span>'}}</td>
      <td style="font-size:.78rem;color:var(--mut)">${{l.address||'—'}}</td>
      <td>${{l.maps_url?`<a href="${{l.maps_url}}" target="_blank">📍 map</a>`:''}}</td>
    </tr>
  `).join('');
}}

function exportCSV(){{
  const cols=['name','rating','reviews','category','phone','website','address','maps_url'];
  const rows=filtered.map(l=>cols.map(c=>`"${{(l[c]||'').toString().replace(/"/g,'""')}}"`).join(','));
  const csv=[cols.join(','),...rows].join('\\n');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([csv],{{type:'text/csv'}}));
  a.download='sg_leads.csv';a.click();
}}

renderStats();render();
</script>
</body>
</html>"""

    path.write_text(html, encoding="utf-8")
    print(f"🌐  Dashboard saved → {path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 58)
    print("  SG Interior Design – Google Maps Lead Scraper")
    print("=" * 58)
    print("\n  A browser window will open — don't close it!")
    print("  The scraper will scroll and collect leads automatically.\n")

    leads = scrape_all()

    if not leads:
        print("⚠️   No leads found. Check your internet connection.")
        return

    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = OUTPUT_DIR / f"sg_leads_{ts}.xlsx"
    html = OUTPUT_DIR / f"sg_leads_dashboard_{ts}.html"

    export_excel(leads, xlsx)
    export_dashboard(leads, html)

    # auto-open dashboard in browser
    import webbrowser
    webbrowser.open(html.resolve().as_uri())

    print()
    print("✨  Done!")
    print(f"   📊  Excel    : {xlsx}")
    print(f"   🌐  Dashboard: {html}")
    print("\n   Open the .html file in any browser to explore your leads.")


if __name__ == "__main__":
    main()
